"""
Kearney Spreadsheet Engine - Brand-Compliant Excel Generation

Generates Excel workbooks that conform to Kearney Design System (KDS) standards.
This module wraps openpyxl to enforce brand compliance automatically.

Usage:
    from core.spreadsheet_engine import KDSSpreadsheet

    wb = KDSSpreadsheet()
    wb.add_cover_sheet(title='Revenue Model', client='Acme Corp')
    wb.add_data_sheet('RawData', dataframe)
    wb.save('exports/model.xlsx')
"""

from pathlib import Path
from typing import List, Dict, Any, Optional, Union
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False


# Kearney Brand Colors
KEARNEY_PURPLE = "7823DC"
KEARNEY_PURPLE_RGB = (120, 35, 220)
DARK_BACKGROUND = "1E1E1E"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
TEXT_DARK = "333333"
TEXT_MEDIUM = "666666"
BORDER_GRAY = "CCCCCC"
INPUT_HIGHLIGHT = "E8D4F8"  # Light purple for input cells


class KDSSpreadsheet:
    """
    Kearney Design System compliant spreadsheet generator.

    Enforces brand standards automatically:
    - Kearney Purple headers
    - Consistent typography (Inter/Arial)
    - No green colors
    - Professional formatting
    """

    def __init__(self):
        """Initialize a new KDS-compliant workbook."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for spreadsheet generation. "
                "Install with: pip install openpyxl"
            )

        self.workbook = openpyxl.Workbook()
        # Remove default sheet - we'll add our own
        self.workbook.remove(self.workbook.active)

        # Define standard styles
        self._setup_styles()

    def _setup_styles(self):
        """Set up standard KDS styles."""
        # Header style - Purple background, white text
        self.header_font = Font(
            name='Arial',  # Inter fallback
            size=11,
            bold=True,
            color=WHITE
        )
        self.header_fill = PatternFill(
            start_color=KEARNEY_PURPLE,
            end_color=KEARNEY_PURPLE,
            fill_type='solid'
        )

        # Body text style
        self.body_font = Font(
            name='Arial',
            size=10,
            color=TEXT_DARK
        )

        # Title style
        self.title_font = Font(
            name='Arial',
            size=16,
            bold=True,
            color=KEARNEY_PURPLE
        )

        # Subtitle style
        self.subtitle_font = Font(
            name='Arial',
            size=12,
            color=TEXT_MEDIUM
        )

        # Input cell style - light purple background
        self.input_fill = PatternFill(
            start_color=INPUT_HIGHLIGHT,
            end_color=INPUT_HIGHLIGHT,
            fill_type='solid'
        )

        # Alternating row style
        self.alt_row_fill = PatternFill(
            start_color=LIGHT_GRAY,
            end_color=LIGHT_GRAY,
            fill_type='solid'
        )

        # Border style
        self.thin_border = Border(
            left=Side(style='thin', color=BORDER_GRAY),
            right=Side(style='thin', color=BORDER_GRAY),
            top=Side(style='thin', color=BORDER_GRAY),
            bottom=Side(style='thin', color=BORDER_GRAY)
        )

        # Alignment
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.left_align = Alignment(horizontal='left', vertical='center')
        self.right_align = Alignment(horizontal='right', vertical='center')

    def add_cover_sheet(
        self,
        title: str,
        client: str = "",
        date: str = "",
        author: str = "",
        version: str = "1.0",
        confidential: bool = True
    ) -> None:
        """
        Add a cover sheet with project metadata.

        Args:
            title: Project/model title
            client: Client name
            date: Date string (defaults to today)
            author: Author name
            version: Version number
            confidential: Whether to mark as confidential
        """
        ws = self.workbook.create_sheet("Cover", 0)

        if not date:
            date = datetime.now().strftime("%Y-%m-%d")

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40

        # Title
        ws.merge_cells('B3:C3')
        ws['B3'] = title
        ws['B3'].font = self.title_font

        # Metadata
        metadata = [
            ("Client:", client),
            ("Date:", date),
            ("Author:", author),
            ("Version:", version),
        ]

        row = 5
        for label, value in metadata:
            ws[f'B{row}'] = label
            ws[f'B{row}'].font = Font(name='Arial', size=11, bold=True, color=TEXT_DARK)
            ws[f'C{row}'] = value
            ws[f'C{row}'].font = self.body_font
            row += 1

        # Confidentiality notice
        if confidential:
            row += 2
            ws.merge_cells(f'B{row}:C{row}')
            ws[f'B{row}'] = "CONFIDENTIAL - For internal use only"
            ws[f'B{row}'].font = Font(name='Arial', size=10, italic=True, color=TEXT_MEDIUM)

    def add_data_sheet(
        self,
        name: str,
        data: Union['pd.DataFrame', List[List[Any]]],
        headers: Optional[List[str]] = None,
        freeze_header: bool = True,
        alternating_rows: bool = True
    ) -> None:
        """
        Add a data sheet with KDS-compliant formatting.

        Args:
            name: Sheet name
            data: DataFrame or list of lists
            headers: Column headers (required if data is list of lists)
            freeze_header: Whether to freeze the header row
            alternating_rows: Whether to use alternating row colors
        """
        ws = self.workbook.create_sheet(name)

        # Handle DataFrame
        if PANDAS_AVAILABLE and isinstance(data, pd.DataFrame):
            headers = list(data.columns)
            rows = data.values.tolist()
        else:
            rows = data
            if headers is None:
                raise ValueError("Headers required when data is not a DataFrame")

        # Write headers
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.body_font
                cell.border = self.thin_border

                # Alternating row colors
                if alternating_rows and row_idx % 2 == 0:
                    cell.fill = self.alt_row_fill

                # Align numbers right, text left
                if isinstance(value, (int, float)):
                    cell.alignment = self.right_align
                else:
                    cell.alignment = self.left_align

        # Freeze header row
        if freeze_header:
            ws.freeze_panes = 'A2'

        # Auto-adjust column widths
        self._auto_column_width(ws)

    def add_summary_sheet(
        self,
        name: str = "Summary",
        metrics: Dict[str, Any] = None
    ) -> None:
        """
        Add a summary sheet with key metrics.

        Args:
            name: Sheet name
            metrics: Dictionary of metric name -> value pairs
        """
        ws = self.workbook.create_sheet(name)

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20

        # Title
        ws.merge_cells('B2:C2')
        ws['B2'] = "Summary"
        ws['B2'].font = self.title_font

        if metrics:
            row = 4
            for metric_name, value in metrics.items():
                ws[f'B{row}'] = metric_name
                ws[f'B{row}'].font = Font(name='Arial', size=11, bold=True, color=TEXT_DARK)
                ws[f'C{row}'] = value
                ws[f'C{row}'].font = self.body_font
                ws[f'C{row}'].alignment = self.right_align
                row += 1

    def add_documentation_sheet(
        self,
        methodology: str = "",
        sources: List[str] = None,
        assumptions: List[str] = None
    ) -> None:
        """
        Add a documentation sheet with methodology and sources.

        Args:
            methodology: Description of methodology
            sources: List of data sources
            assumptions: List of key assumptions
        """
        ws = self.workbook.create_sheet("Documentation")

        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 80

        row = 2

        # Methodology section
        ws[f'B{row}'] = "Methodology"
        ws[f'B{row}'].font = Font(name='Arial', size=14, bold=True, color=KEARNEY_PURPLE)
        row += 1
        if methodology:
            ws[f'B{row}'] = methodology
            ws[f'B{row}'].font = self.body_font
            ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 2

        # Sources section
        ws[f'B{row}'] = "Data Sources"
        ws[f'B{row}'].font = Font(name='Arial', size=14, bold=True, color=KEARNEY_PURPLE)
        row += 1
        if sources:
            for source in sources:
                ws[f'B{row}'] = f"- {source}"
                ws[f'B{row}'].font = self.body_font
                row += 1
        row += 1

        # Assumptions section
        ws[f'B{row}'] = "Key Assumptions"
        ws[f'B{row}'].font = Font(name='Arial', size=14, bold=True, color=KEARNEY_PURPLE)
        row += 1
        if assumptions:
            for assumption in assumptions:
                ws[f'B{row}'] = f"- {assumption}"
                ws[f'B{row}'].font = self.body_font
                row += 1

    def mark_input_cells(self, sheet_name: str, cell_ranges: List[str]) -> None:
        """
        Mark cells as input cells (light purple highlight).

        Args:
            sheet_name: Name of the sheet
            cell_ranges: List of cell references (e.g., ['B3', 'B4:B10'])
        """
        ws = self.workbook[sheet_name]

        for cell_range in cell_ranges:
            if ':' in cell_range:
                # Range of cells
                for row in ws[cell_range]:
                    for cell in row:
                        cell.fill = self.input_fill
            else:
                # Single cell
                ws[cell_range].fill = self.input_fill

    def _auto_column_width(self, worksheet, min_width: int = 10, max_width: int = 50):
        """Auto-adjust column widths based on content."""
        for column_cells in worksheet.columns:
            length = max(
                len(str(cell.value or "")) for cell in column_cells
            )
            length = min(max(length + 2, min_width), max_width)
            worksheet.column_dimensions[column_cells[0].column_letter].width = length

    def save(self, path: Union[str, Path]) -> Path:
        """
        Save the workbook to file.

        Args:
            path: Output file path

        Returns:
            Path to the saved file
        """
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(path)
        return path


def create_data_table(
    data: Union['pd.DataFrame', List[List[Any]]],
    headers: List[str] = None,
    output_path: Union[str, Path] = None,
    title: str = "Data Export",
    client: str = "",
    include_cover: bool = True
) -> Path:
    """
    Convenience function to create a simple data table workbook.

    Args:
        data: DataFrame or list of lists
        headers: Column headers (required if data is list of lists)
        output_path: Where to save the file
        title: Title for cover sheet
        client: Client name for cover sheet
        include_cover: Whether to include a cover sheet

    Returns:
        Path to the saved file
    """
    wb = KDSSpreadsheet()

    if include_cover:
        wb.add_cover_sheet(title=title, client=client)

    wb.add_data_sheet("Data", data, headers=headers)

    if output_path is None:
        output_path = f"outputs/{title.lower().replace(' ', '_')}.xlsx"

    return wb.save(output_path)
